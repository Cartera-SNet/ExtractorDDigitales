# -*- coding: utf-8 -*-
"""
Extractor de Documentos Digitales — v1 (modo local / pruebas)
=============================================================

Qué hace esta primera versión:
  1. Recibe uno o varios PDF (los que hoy salen de la reclamación:
     FURIPS + factura + epicrisis + anexos + fotos de documentos).
  2. Identifica dentro de cada PDF las páginas que corresponden a:
        - SIRAS
        - CÉDULA (o Permiso por Protección Temporal / doc. migrante)
        - TARJETA DE PROPIEDAD (licencia de tránsito)
  3. Arma un PDF nuevo, nombrado con el número de factura, que trae
     SOLO esos documentos, en el orden: SIRAS, CÉDULA, TARJETA.
  4. Genera un Excel con el resultado de cada archivo procesado,
     indicando si quedó completo o qué documento faltó.

Lo que falta para la versión final (a propósito, no está en este v1):
  - Conexión a la API/base de datos real para traer los PDF por IPS/sede
    en vez de subirlos manualmente.
  - Cruce con el mapeo de IPS/NIT (config/mapeo_ips.xlsx ya viene incluido
    como referencia para la siguiente etapa).

Arranque: doble clic en iniciar.bat, luego abrir http://localhost:5057
"""

import os
import io
import csv
import json
import re
import base64
import shutil
import fitz
import threading
import time
import zipfile
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for

import clasificador as clf
import auth
import checklist_config as chk
import progreso as prog

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except Exception:
    EXCEL_OK = False


BASE_DIR = Path(__file__).resolve().parent
UPLOADS_DIR = BASE_DIR / "uploads"

# Se detecta sola si esto corre en la nube (Railway pone estas variables
# automáticamente) o en una máquina local (iniciar.bat) — así el
# encabezado de la interfaz deja de decir "pruebas locales" a propósito
# cuando el bot ya está sirviendo de verdad en la web, sin tener que
# acordarse de cambiar nada a mano en cada despliegue.
ETIQUETA_ENTORNO = ("v1 · en línea" if (os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("PORT"))
                     else "v1 · pruebas locales")
OUTPUTS_DIR = BASE_DIR / "outputs"
UPLOADS_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 300 * 1024 * 1024  # 300 MB por lote
app.secret_key = os.environ.get("SECRET_KEY", "bot-soat-2026-key-local")
app.permanent_session_lifetime = timedelta(hours=8)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"


# ─────────────────────────────────────────────────────────────
# Estado del lote actual (para el log en vivo + barra de progreso).
# El procesamiento corre en un hilo aparte; el navegador va preguntando
# "¿cómo vas?" cada cierto tiempo (polling) a /api/estado. Sirve tanto
# para lotes chicos de prueba como para lotes grandes (100-300+ casos),
# donde ver el avance en vivo importa bastante más.
# ─────────────────────────────────────────────────────────────

ESTADO_LOCK = threading.Lock()
EVENTO_DETENER = threading.Event()


def _estado_base():
    """Estado 'de fábrica' — se usa tanto al arrancar el servidor como
    para resetear TODO después de 'Borrar progreso', para no dejar
    ningún campo residual de una corrida anterior (ver máquina de
    estados del botón principal, sección de abajo)."""
    return {
        "running": False,
        "finished": False,
        "error": None,
        "logs": [],
        "stats": {"total": 0, "ok": 0, "err": 0},
        "lote_id": None,
        "resumenes": [],
        "pdfs": [],
        "tiene_excel": False,
        "rutas_sugeridas": [],  # [{"caso":.., "ruta":..}] — dónde SÍ se encontró cada caso
        "resumen_errores": [],  # [{"referencia":.., "mensaje":..}] — errores en lenguaje simple
        "urls_consulta": [],  # [{"caso":.., "ruta":.., "url":..}] — URL exacta llamada por cada intento, para probar a mano
        "detenido": False,  # si el usuario le dio "Detener" a mitad del lote
        "hilo": None,  # referencia real al Thread que está corriendo (o corrió) el lote actual
        "limpiando": False,  # evita que 2 clics seguidos en "Borrar progreso" se pisen entre sí
        "modo": None,  # "casos" | "subida" -- cuál panel disparó el lote actual/último
    }


ESTADO_ACTUAL = _estado_base()


# ─────────────────────────────────────────────────────────────
# Máquina de estados del botón principal
# ─────────────────────────────────────────────────────────────
# El FRONTEND nunca decide el estado por su cuenta — siempre lo pregunta
# aquí. La verdad no es "una bandera booleana que alguien puso en True",
# sino si el HILO realmente sigue vivo (`Thread.is_alive()`). Si el
# proceso se cayó por una excepción no controlada, un apagón, o
# cualquier causa externa, sin pasar por el código que limpia
# "running", este chequeo lo detecta solo y se autocorrige.
#
#   ejecutando -> el hilo del lote actual sigue vivo de verdad
#   pausado    -> no hay hilo vivo, pero existe progreso.json sin
#                 terminar (en_proceso/detenido) para el último lote
#   listo      -> no hay hilo vivo ni progreso pendiente

def _hilo_vivo_sin_lock() -> bool:
    """Igual que `_hilo_activo()` pero SIN tomar el lock — solo para
    usar desde código que YA está dentro de un `with ESTADO_LOCK:`
    (threading.Lock no es reentrante; tomarlo dos veces desde el mismo
    hilo se traba para siempre)."""
    hilo = ESTADO_ACTUAL.get("hilo")
    return bool(hilo and hilo.is_alive())


def _hilo_activo() -> bool:
    with ESTADO_LOCK:
        return _hilo_vivo_sin_lock()


def _sanear_estado():
    """Se llama al INICIO de cualquier endpoint que lea o cambie el
    estado del proceso. Si `running=True` mentía (el hilo ya no existe
    porque el proceso se cayó sin avisar), se corrige aquí — así ningún
    endpoint puede quedar confundido por una bandera vieja."""
    se_corrigio = False
    with ESTADO_LOCK:
        hilo = ESTADO_ACTUAL.get("hilo")
        if ESTADO_ACTUAL["running"] and not (hilo and hilo.is_alive()):
            ESTADO_ACTUAL["running"] = False
            ESTADO_ACTUAL["finished"] = True
            if not ESTADO_ACTUAL.get("error"):
                ESTADO_ACTUAL["error"] = "El proceso se interrumpió sin avisar (excepción no controlada o cierre inesperado)."
            se_corrigio = True
    if se_corrigio:  # el _log() de acá abajo toma el lock por su cuenta -- no puede ir dentro del "with" de arriba
        _log("Se detectó que el proceso ya no estaba realmente corriendo — el estado se corrigió solo.", level="warn")


def _estado_boton(sesion_datos=None):
    """Calcula el estado real del botón principal: 'ejecutando',
    'pausado' o 'listo'. `sesion_datos` es opcional -- si se manda, se
    usa para revisar si hay progreso pendiente de ESTE usuario/empresa
    aunque el navegador se haya recargado y no tenga el lote_id a mano."""
    _sanear_estado()
    if _hilo_activo():
        return "ejecutando"
    with ESTADO_LOCK:
        lote_actual = ESTADO_ACTUAL.get("lote_id")
    if lote_actual and prog.existe_progreso_activo(lote_actual):
        return "pausado"
    if sesion_datos:
        pendiente = prog.buscar_progreso_pendiente(sesion_datos.get("usuario", "?"), sesion_datos.get("empresa", "?"))
        if pendiente:
            return "pausado"
    return "listo"


def _solo_digitos(valor) -> str:
    """Normaliza un número de caso eliminando TODO lo que no sea dígito
    (espacios, guiones, puntos, barras, paréntesis, corchetes, texto
    como 'Caso:' o 'N°', etc.) — así '641 757', '641-757', '(641757)',
    'N° 641757' y similares terminan siendo todos '641757'. Se usa como
    único punto de normalización antes de guardar/mandar un NoCaso a la
    API, para no depender de que el archivo Excel o la digitación manual
    vengan perfectamente limpios."""
    return re.sub(r"[^0-9]", "", str(valor or ""))


def _log(msg, level="info"):
    """Agrega una línea al log en vivo (con hora), pensado para leerse
    desde /api/estado casi en tiempo real mientras corre el lote."""
    with ESTADO_LOCK:
        ESTADO_ACTUAL["logs"].append({
            "ts": datetime.now().strftime("%H:%M:%S"),
            "level": level,
            "msg": msg,
        })


def _sumar_stat(campo):
    with ESTADO_LOCK:
        ESTADO_ACTUAL["stats"][campo] = ESTADO_ACTUAL["stats"].get(campo, 0) + 1


def _registrar_ruta_sugerida(no_caso, nombre_ruta):
    """Se llama cada vez que un caso SÍ se encuentra en una ruta — con el
    tiempo, esta lista sirve para saber cuál ruta conviene dejar marcada
    por defecto para cada tipo de caso."""
    with ESTADO_LOCK:
        ESTADO_ACTUAL["rutas_sugeridas"].append({"caso": no_caso, "ruta": nombre_ruta})


def _registrar_url_consulta(no_caso, nombre_ruta, url):
    """Guarda la URL EXACTA que se llamó para buscar un caso en una ruta
    — se registra en TODOS los intentos (encontrado o 404), justamente
    porque donde más ayuda es cuando dio 404 y se quiere probar a mano
    en el navegador si de verdad no está ahí o si es un problema de
    parámetros. Vive en su propia pestaña ('URLs de búsqueda'), separada
    de 'Rutas sugeridas' (que solo muestra dónde SÍ se encontró)."""
    with ESTADO_LOCK:
        ESTADO_ACTUAL["urls_consulta"].append({"caso": no_caso, "ruta": nombre_ruta, "url": url})


def _mensaje_simple(error, limite=140):
    """Recorta un mensaje de error técnico a algo cortico y legible para
    cualquier persona — se queda con la primera oración (donde ya suelen
    estar los mensajes de este proyecto redactados para que se entiendan
    solos) y descarta el detalle técnico que viene después."""
    texto = str(error).strip()
    primera = texto.split(". ")[0]
    if len(primera) > limite:
        primera = primera[:limite].rstrip() + "..."
    return primera if primera.endswith((".", "...")) else primera + "."


def _registrar_error_resumen(referencia, error):
    with ESTADO_LOCK:
        ESTADO_ACTUAL["resumen_errores"].append({
            "referencia": referencia,
            "mensaje": _mensaje_simple(error),
        })


# ─────────────────────────────────────────────────────────────
# Procesamiento de un lote de PDFs
# ─────────────────────────────────────────────────────────────

def _limpiar_carpeta(carpeta: Path):
    for item in carpeta.iterdir():
        if item.is_file():
            item.unlink()
        elif item.is_dir():
            shutil.rmtree(item)


# Cuántos archivos se procesan a la vez en un mismo lote. Cada archivo ya
# reparte sus propias páginas entre varios hilos dentro de clasificador.py
# (clf.HILOS_OCR), así que aquí se usa un número más chico para no
# disparar demasiados procesos de Tesseract al mismo tiempo — en una
# máquina con muchos núcleos (pensando en lotes de 100-300+ casos) esto
# igual deja bastante paralelismo real.
HILOS_ARCHIVOS = min(2, max(1, (os.cpu_count() or 4) // 4))


def _log_detalle_resultado(nombre_original, paginas, resumen):
    """Registra en el log, línea por línea, qué encontró el bot para este
    archivo — no solo el estado final, sino documento por documento, para
    poder ver de un vistazo qué pasó sin tener que abrir el Excel."""
    total_paginas = len(paginas)
    paginas_ocr = sum(1 for p in paginas if p.ocr_usado)
    _log(f"{nombre_original}: {total_paginas} página(s) revisada(s) "
         f"({paginas_ocr} por OCR, el resto ya traían texto digital)")

    # Puntajes por página escaneada: ayuda a diagnosticar cuando un
    # documento "no aparece" — si el puntaje quedó cerca del mínimo
    # (UMBRAL_PUNTAJE) puede ser que el OCR de esa máquina lea un poco
    # peor el texto y no llegue a tiempo; con este detalle se puede
    # ajustar sin tener que pedir capturas de pantalla.
    for p in paginas:
        if p.ocr_usado:
            pts = p.puntajes
            _log(f"{nombre_original}: pág. {p.numero} (OCR) → puntajes: "
                 f"SIRAS={pts.get('SIRAS',0)}, CEDULA={pts.get('CEDULA',0)}, "
                 f"TARJETA={pts.get('TARJETA_PROPIEDAD',0)} (mínimo para contar: {clf.UMBRAL_PUNTAJE})")
            # Se registra exactamente qué pasó al intentar separar —
            # sin esto, cuando NO se separa, no había forma de saber en
            # qué punto se quedó corto el intento (se muestra siempre que
            # haya diagnóstico, no solo con 2+ categorías, porque a veces
            # el intento se hace igual con una sola categoría detectada).
            if p.diagnostico_separacion:
                nivel_log = "info" if p.dividir else "warn"
                _log(f"{nombre_original}: pág. {p.numero} — diagnóstico de separación: "
                     f"{p.diagnostico_separacion}", level=nivel_log)

    def _linea_doc(nombre_doc, encontrado, paginas_txt):
        if encontrado == "Sí":
            _log(f"{nombre_original}: {nombre_doc} → encontrado (pág. {paginas_txt})", level="ok")
        else:
            _log(f"{nombre_original}: {nombre_doc} → NO encontrado", level="warn")

    _linea_doc("SIRAS", resumen.get("SIRAS"), resumen.get("Pág. SIRAS", ""))
    _linea_doc("Cédula/Migrante", resumen.get("Cédula/Migrante"), resumen.get("Pág. Cédula", ""))
    _linea_doc("Tarjeta de propiedad", resumen.get("Tarjeta Propiedad"), resumen.get("Pág. Tarjeta", ""))

    divididas = [p.numero for p in paginas if p.dividir]
    if divididas:
        _log(f"{nombre_original}: se separaron documentos que venían pegados en la "
             f"misma foto, en la(s) página(s) {', '.join(map(str, divididas))}", level="info")


def _procesar_un_archivo(item, categorias_permitidas=None):
    nombre_original, ruta = item
    _log(f"Procesando {nombre_original}...")
    try:
        paginas = clf.clasificar_pdf(str(ruta), debe_detener=EVENTO_DETENER.is_set)
        if EVENTO_DETENER.is_set():
            _log(f"{nombre_original}: detención solicitada — se sigue con lo que ya se alcanzó "
                 f"a analizar de este archivo (puede quedar INCOMPLETO por eso, no porque falten "
                 f"documentos de verdad).", level="warn")
        factura = clf.extraer_numero_factura(paginas, nombre_original)
        pdf_bytes, por_categoria, orden = clf.construir_pdf_unificado(
            str(ruta), paginas, categorias_permitidas=categorias_permitidas
        )
        resumen = clf.resumir_resultado(
            nombre_original, factura, por_categoria, paginas
        )

        _log_detalle_resultado(nombre_original, paginas, resumen)

        if orden:
            nombre_salida = f"{factura}.pdf"
            resumen["Archivo generado"] = nombre_salida
            generado = (nombre_salida, pdf_bytes)
        else:
            resumen["Archivo generado"] = "(no generado - 0 páginas identificadas)"
            generado = None

        estado = resumen.get("Estado", "")
        nivel = "ok" if estado == "COMPLETO" else ("warn" if estado == "INCOMPLETO" else "error")
        _log(f"{nombre_original} → {estado} (factura {factura})", level=nivel)
        _sumar_stat("ok" if estado in ("COMPLETO", "INCOMPLETO") else "err")
        return resumen, generado

    except Exception as e:
        traceback.print_exc()
        _log(f"Error procesando {nombre_original}: {e}", level="error")
        _sumar_stat("err")
        _registrar_error_resumen(nombre_original, f"Error procesando el archivo: {e}")
        resumen_error = {
            "Archivo original": nombre_original,
            "No. Factura": "",
            "SIRAS": "?",
            "Pág. SIRAS": "",
            "Cédula/Migrante": "?",
            "Pág. Cédula": "",
            "Tarjeta Propiedad": "?",
            "Pág. Tarjeta": "",
            "Estado": "ERROR",
            "Observaciones": f"Error procesando el archivo: {e}",
            "Archivo generado": "",
        }
        return resumen_error, None


def procesar_lote(rutas_pdf):
    """rutas_pdf: lista de (nombre_original, ruta_en_disco)

    Devuelve (lista_resumenes, lista_pdfs_generados[(nombre_archivo, bytes)])

    Los archivos del lote se procesan con algo de paralelismo (ver
    HILOS_ARCHIVOS) — importante de cara a lotes grandes (100-300+
    casos), donde procesar uno por uno sería mucho más lento. El orden de
    los resultados en pantalla se mantiene igual al de subida, aunque el
    procesamiento interno no sea estrictamente en ese orden.
    """
    resumenes_por_indice = {}
    pdfs_generados = []
    detenido = False

    if len(rutas_pdf) > 1:
        ex = ThreadPoolExecutor(max_workers=HILOS_ARCHIVOS)
        futuros = {
            ex.submit(_procesar_un_archivo, item): idx
            for idx, item in enumerate(rutas_pdf)
        }
        try:
            for futuro in as_completed(futuros):
                idx = futuros[futuro]
                resumen, generado = futuro.result()
                resumenes_por_indice[idx] = resumen
                if generado:
                    pdfs_generados.append(generado)
                if EVENTO_DETENER.is_set():
                    detenido = True
                    _log("Detención solicitada — se completan los archivos que ya estaban en "
                         "proceso y se arma el resultado con lo que haya listo hasta ahora.", level="warn")
                    break
        finally:
            # cancel_futures=True: los archivos que ni siquiera habían
            # empezado a procesarse se descartan; los que ya estaban
            # corriendo en un hilo no se pueden interrumpir a mitad de
            # camino, pero se dejan terminar solos (nunca se cortan a la
            # mitad, para no dejar un PDF a medio generar).
            ex.shutdown(wait=True, cancel_futures=True)
    else:
        for idx, item in enumerate(rutas_pdf):
            if EVENTO_DETENER.is_set():
                detenido = True
                _log("Detención solicitada antes de procesar el siguiente archivo.", level="warn")
                break
            resumen, generado = _procesar_un_archivo(item)
            resumenes_por_indice[idx] = resumen
            if generado:
                pdfs_generados.append(generado)

    # Si se detuvo a mitad de camino, el resultado solo incluye los
    # índices que sí alcanzaron a procesarse (no todos los originales).
    resumenes = [resumenes_por_indice[i] for i in range(len(rutas_pdf)) if i in resumenes_por_indice]
    return resumenes, pdfs_generados, detenido


# ─────────────────────────────────────────────────────────────
# Excel de resultados
# ─────────────────────────────────────────────────────────────

COLUMNAS = [
    "Archivo original", "No. Factura", "SIRAS", "Pág. SIRAS",
    "Cédula/Migrante", "Pág. Cédula", "Tarjeta Propiedad", "Pág. Tarjeta",
    "Archivo generado", "Estado", "Observaciones",
]

COLOR_OK = "C6EFCE"
COLOR_INCOMPLETO = "FFEB9C"
COLOR_ERROR = "FFC7CE"


def generar_excel(resumenes):
    if not EXCEL_OK:
        raise RuntimeError("openpyxl no está instalado")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    fill_map = {
        "COMPLETO": PatternFill(start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid"),
        "INCOMPLETO": PatternFill(start_color=COLOR_INCOMPLETO, end_color=COLOR_INCOMPLETO, fill_type="solid"),
        "SIN DOCUMENTOS": PatternFill(start_color=COLOR_ERROR, end_color=COLOR_ERROR, fill_type="solid"),
        "ERROR": PatternFill(start_color=COLOR_ERROR, end_color=COLOR_ERROR, fill_type="solid"),
    }

    for row_idx, resumen in enumerate(resumenes, start=2):
        estado = resumen.get("Estado", "")
        fill = fill_map.get(estado)
        for col_idx, col_name in enumerate(COLUMNAS, start=1):
            valor = resumen.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if fill:
                cell.fill = fill

    anchos = [22, 14, 8, 10, 16, 10, 16, 10, 22, 14, 45]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────
# Autenticación (Servidor + Usuario + Contraseña + Empresa)
# ─────────────────────────────────────────────────────────────

@app.route("/login")
def login_page():
    if session.get("autenticado"):
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/api/servidores")
def api_servidores():
    """Pública — la usa la pantalla de login para llenar el combo de
    servidores antes de que el usuario se autentique."""
    servidores = auth.get_servidores_api()
    data = {}
    for srv in servidores:
        nombre = (srv.get("Descripcion") or "").strip()
        if nombre:
            data[nombre] = {
                "ip": srv.get("Ip_Conexion", ""),
                "puerto": str(srv.get("Puerto", "3306")),
            }
    return jsonify(data)


@app.route("/api/auth/conectar", methods=["POST"])
def api_auth_conectar():
    """Paso 1 del login: valida usuario/clave contra el servidor elegido
    y devuelve las empresas a las que tiene acceso (todavía sin crear
    sesión — eso pasa en el paso 2, cuando ya eligió la empresa)."""
    ip_solicitante = auth.ip_solicitante()
    ok, msg = auth.chequear_rate(ip_solicitante)
    if not ok:
        return jsonify({"error": msg}), 429

    data = request.get_json() or {}
    usuario = data.get("usuario", "").strip()
    password = data.get("password", "").strip()
    servidor = data.get("servidor", "").strip()

    if not all([usuario, password, servidor]):
        return jsonify({"error": "Completa todos los campos"}), 400

    servidores = auth.get_servidores_api()
    srv = next((s for s in servidores if (s.get("Descripcion") or "").strip() == servidor), None)
    if not srv:
        return jsonify({"error": f"Servidor '{servidor}' no encontrado"}), 404

    ip_srv = srv.get("Ip_Conexion", "")
    puerto = str(srv.get("Puerto", "3306"))

    empresas_raw = auth.get_empresas_api(ip_srv, puerto, usuario, password)
    if not empresas_raw:
        auth.marcar_fallo(ip_solicitante)
        return jsonify({"error": "Credenciales inválidas o sin empresas asignadas."}), 401

    empresas = [
        {"cod": e.get("Empresa", ""), "nombre": e.get("Nombre_Empresa", ""), "sede": e.get("NombreComercial", "")}
        for e in empresas_raw if e.get("Empresa")
    ]
    auth.marcar_exito(ip_solicitante)
    return jsonify({"success": True, "empresas": empresas, "ip": ip_srv, "puerto": puerto})


@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    """Paso 2 del login: ya con la empresa elegida, crea la sesión."""
    ip_solicitante = auth.ip_solicitante()
    ok, msg = auth.chequear_rate(ip_solicitante)
    if not ok:
        return jsonify({"error": msg}), 429

    data = request.get_json() or {}
    usuario = data.get("usuario", "").strip()
    password = data.get("password", "").strip()
    servidor = data.get("servidor", "").strip()
    empresa = data.get("empresa", "").strip()
    empresas = data.get("empresas", [])

    if not all([usuario, password, servidor, empresa]):
        return jsonify({"error": "Datos incompletos"}), 400

    servidores = auth.get_servidores_api()
    srv = next((s for s in servidores if (s.get("Descripcion") or "").strip() == servidor), None)
    if not srv:
        return jsonify({"error": "Servidor no encontrado"}), 404

    session.permanent = True
    session["autenticado"] = True
    session["usuario"] = usuario
    session["password"] = password
    session["servidor"] = servidor
    session["ip"] = srv.get("Ip_Conexion", "")
    session["puerto"] = str(srv.get("Puerto", "3306"))
    session["empresa"] = empresa
    session["empresas"] = empresas
    return jsonify({"success": True})


@app.route("/api/auth/cambiar-empresa", methods=["POST"])
@auth.login_requerido
def api_auth_cambiar_empresa():
    data = request.get_json(force=True) or {}
    empresa_cod = data.get("empresa", "").strip()
    empresas = session.get("empresas", [])
    if not any(e.get("cod") == empresa_cod for e in empresas):
        return jsonify({"error": "Empresa no válida para este usuario"}), 403
    session["empresa"] = empresa_cod
    return jsonify({"success": True, "empresa": empresa_cod})


@app.route("/api/auth/logout", methods=["POST"])
def api_auth_logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/auth/me")
def api_auth_me():
    if not session.get("autenticado"):
        return jsonify({"autenticado": False}), 401
    empresa_cod = session.get("empresa", "")
    empresas = session.get("empresas", [])
    empresa_nombre = empresa_cod
    for e in empresas:
        if e.get("cod") == empresa_cod:
            sede = f" · {e['sede']}" if e.get("sede") else ""
            empresa_nombre = f"{empresa_cod} — {e.get('nombre', '')}{sede}"
            break
    return jsonify({
        "autenticado": True,
        "usuario": session.get("usuario", ""),
        "servidor": session.get("servidor", ""),
        "empresa": empresa_cod,
        "empresa_nombre": empresa_nombre,
        "empresas": empresas,
    })


# ─────────────────────────────────────────────────────────────
# Rutas Flask
# ─────────────────────────────────────────────────────────────

@app.route("/")
@auth.login_requerido
def index():
    sesion = auth.sesion_actual()
    items_paciente, error_paciente = auth.get_lista_tipo_documentos_api(
        sesion["ip"], sesion["puerto"], sesion["usuario"], sesion["password"], "AP"
    )
    items_caso, error_caso = auth.get_lista_tipo_documentos_api(
        sesion["ip"], sesion["puerto"], sesion["usuario"], sesion["password"], "AC"
    )
    items_rutas, error_rutas = auth.get_centro_digital_api(
        sesion["ip"], sesion["puerto"], sesion["usuario"], sesion["password"], sesion["empresa"]
    )

    grupo_paciente = chk.construir_grupo(items_paciente, "paciente")
    grupo_caso = chk.construir_grupo(items_caso, "caso")
    grupo_rutas = chk.construir_grupo(items_rutas, "ruta")

    return render_template(
        "index.html",
        checklist={"paciente": grupo_paciente, "caso": grupo_caso, "rutas": grupo_rutas},
        checklist_default=chk.ids_por_defecto(grupo_paciente, grupo_caso, grupo_rutas),
        checklist_errores={"paciente": error_paciente, "caso": error_caso, "rutas": error_rutas},
        observacion_rutas=chk.OBSERVACION_RUTAS,
        etiqueta_entorno=ETIQUETA_ENTORNO,
    )


def _ejecutar_lote_en_hilo(lote_id, carpeta_lote, rutas_pdf):
    """Corre todo el procesamiento del lote en un hilo aparte, para que
    el navegador pueda ir preguntando el avance (log en vivo) mientras
    tanto en vez de quedar esperando una sola respuesta larga."""
    _log(f"Iniciando lote con {len(rutas_pdf)} archivo(s)...")
    resumenes, pdfs_generados, error, detenido = [], [], None, False
    try:
        resumenes, pdfs_generados, detenido = procesar_lote(rutas_pdf)
    except Exception as e:
        traceback.print_exc()
        error = str(e)
        _log(f"Error inesperado procesando el lote: {e}", level="error")

    carpeta_salida = OUTPUTS_DIR / lote_id
    carpeta_salida.mkdir(parents=True, exist_ok=True)
    for nombre_pdf, contenido in pdfs_generados:
        with open(carpeta_salida / nombre_pdf, "wb") as fh:
            fh.write(contenido)

    excel_bytes = None
    if EXCEL_OK and resumenes:
        try:
            excel_bytes = generar_excel(resumenes)
            with open(carpeta_salida / "Informe_resultado.xlsx", "wb") as fh:
                fh.write(excel_bytes)
        except Exception as e:
            _log(f"No se pudo generar el Excel: {e}", level="warn")

    shutil.rmtree(carpeta_lote, ignore_errors=True)

    if detenido:
        _log(f"Lote detenido por el usuario — se procesaron {len(resumenes)} de {len(rutas_pdf)} "
             f"archivo(s) antes de parar. El ZIP/Excel ya están listos con eso.", level="warn")
    else:
        _log("Lote terminado." if not error else "Lote terminado con errores.",
             level="ok" if not error else "error")

    with ESTADO_LOCK:
        ESTADO_ACTUAL["running"] = False
        ESTADO_ACTUAL["finished"] = True
        ESTADO_ACTUAL["error"] = error
        ESTADO_ACTUAL["detenido"] = detenido
        ESTADO_ACTUAL["resumenes"] = resumenes
        ESTADO_ACTUAL["pdfs"] = [n for n, _ in pdfs_generados]
        ESTADO_ACTUAL["tiene_excel"] = bool(excel_bytes)


@app.route("/procesar", methods=["POST"])
@auth.login_requerido
def procesar():
    _sanear_estado()
    with ESTADO_LOCK:
        if _hilo_vivo_sin_lock():
            return jsonify({"ok": False, "error": "Ya hay un lote en proceso, espera a que termine."}), 409
        # Se reserva el turno DE UNA, antes de tocar archivos — si no se
        # hace así, dos clics casi al mismo tiempo (doble clic, doble
        # submit) pueden colarse los dos antes de que se alcance a marcar
        # "running=True" más abajo, y arrancan dos lotes a la vez.
        ESTADO_ACTUAL["running"] = True

    EVENTO_DETENER.clear()
    archivos = request.files.getlist("pdfs")
    if not archivos:
        with ESTADO_LOCK:
            ESTADO_ACTUAL["running"] = False
        return jsonify({"ok": False, "error": "No se recibió ningún PDF."}), 400

    lote_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    carpeta_lote = UPLOADS_DIR / lote_id
    carpeta_lote.mkdir(parents=True, exist_ok=True)

    rutas_pdf = []
    for f in archivos:
        if not f.filename.lower().endswith(".pdf"):
            continue
        destino = carpeta_lote / f.filename
        f.save(destino)
        rutas_pdf.append((f.filename, destino))

    if not rutas_pdf:
        with ESTADO_LOCK:
            ESTADO_ACTUAL["running"] = False
        return jsonify({"ok": False, "error": "Ninguno de los archivos es un PDF válido."}), 400

    with ESTADO_LOCK:
        ESTADO_ACTUAL.update({
            "running": True,
            "finished": False,
            "error": None,
            "logs": [],
            "stats": {"total": len(rutas_pdf), "ok": 0, "err": 0},
            "lote_id": lote_id,
            "resumenes": [],
            "pdfs": [],
            "tiene_excel": False,
            "rutas_sugeridas": [],
            "resumen_errores": [],
            "urls_consulta": [],
            "modo": "subida",
        })

    hilo = threading.Thread(
        target=_ejecutar_lote_en_hilo,
        args=(lote_id, carpeta_lote, rutas_pdf),
        daemon=True,
    )
    with ESTADO_LOCK:
        ESTADO_ACTUAL["hilo"] = hilo
    hilo.start()

    return jsonify({"ok": True, "lote_id": lote_id, "total": len(rutas_pdf), "estado_boton": "ejecutando"})


@app.route("/api/detener", methods=["POST"])
@auth.login_requerido
def api_detener():
    """Detiene el lote/búsqueda en curso lo antes posible: cancela los
    archivos/casos que ni siquiera habían empezado a procesarse, deja
    terminar los que ya estaban en un hilo corriendo (para no cortar un
    PDF a medio generar), y arma el ZIP/Excel con lo que haya quedado
    listo hasta ese momento — igual que un lote que termina normal, solo
    que con menos archivos de los que se pidieron al principio."""
    _sanear_estado()
    if not _hilo_activo():
        return jsonify({"ok": False, "error": "No hay ningún lote corriendo en este momento."}), 409
    EVENTO_DETENER.set()
    _log("Deteniendo a pedido del usuario...", level="warn")
    return jsonify({"ok": True})


@app.route("/api/limpiar-caso", methods=["POST"])
@auth.login_requerido
def api_limpiar_caso():
    """Botón 'Borrar progreso': reinicio real y completo.

    Sigue un orden seguro (no una limpieza 'a medias' que deje al
    backend pensando que sigue habiendo algo corriendo):
      1) si el lote activo es justo el de estos casos, se le pide que
         se detenga y se ESPERA (join) a que el hilo de verdad termine
         — nunca se asume que terminó solo porque se puso la bandera;
      2) recién ahí se borra progreso.json + carpetas de trabajo;
      3) se reinicia TODO el estado global a los valores de fábrica
         (ver `_estado_base()`) — nada de dejar `lote_id`, `resumenes`,
         `detenido`, etc. de la corrida anterior colgando por ahí;
      4) se responde con el estado ya limpio, para que el navegador no
         tenga que adivinar ni esperar a la próxima consulta de polling.

    Un flag `limpiando` (con el mismo ESTADO_LOCK) evita que dos clics
    seguidos en el botón, o un doble envío accidental, disparen dos
    limpiezas a la vez pisándose una a la otra.
    """
    with ESTADO_LOCK:
        if ESTADO_ACTUAL.get("limpiando"):
            return jsonify({"ok": False, "error": "Ya se está limpiando — espera un momento."}), 409
        ESTADO_ACTUAL["limpiando"] = True

    try:
        _sanear_estado()
        data = request.get_json() or {}
        casos = data.get("casos", [])
        if not casos:
            return jsonify({"ok": False, "error": "No se recibió ningún caso para identificar qué limpiar."}), 400
        casos = [{"NoCaso": _solo_digitos(c.get("NoCaso"))} for c in casos if _solo_digitos(c.get("NoCaso"))]
        if not casos:
            return jsonify({"ok": False, "error": "Ninguno de los casos recibidos tiene un número válido."}), 400

        sesion_datos = auth.sesion_actual()
        clave_lote = prog.calcular_clave_lote(casos, sesion_datos.get("usuario", "?"), sesion_datos.get("empresa", "?"))

        with ESTADO_LOCK:
            es_el_lote_activo = ESTADO_ACTUAL.get("lote_id") == clave_lote
            hilo = ESTADO_ACTUAL.get("hilo")

        if es_el_lote_activo and hilo and hilo.is_alive():
            _log("Deteniendo el proceso activo para poder borrar su progreso...", level="warn")
            EVENTO_DETENER.set()
            hilo.join(timeout=90)  # generoso: deja terminar el caso que ya estaba en curso, nunca lo corta a la fuerza
            if hilo.is_alive():
                # No se puede matar un hilo de Python a la fuerza sin
                # arriesgar un estado inconsistente (un PDF a medio
                # escribir, un archivo de progreso a medio guardar) — se
                # prefiere avisar y que el usuario reintente en vez de
                # forzar algo peligroso.
                return jsonify({
                    "ok": False,
                    "error": "El proceso todavía se está deteniendo (un caso en curso no terminó a tiempo) — intenta de nuevo en unos segundos.",
                }), 409
            EVENTO_DETENER.clear()

        existia = prog.cargar_progreso(clave_lote) is not None
        prog.limpiar_todo(clave_lote, carpetas_extra=[UPLOADS_DIR / clave_lote, OUTPUTS_DIR / clave_lote])

        with ESTADO_LOCK:
            if es_el_lote_activo:
                # Reset TOTAL -- este era el lote que estaba trackeado en
                # el estado global, así que no debe quedar NADA residual
                # de él (ni "detenido", ni "finished", ni resumenes viejos).
                nuevo = _estado_base()
                nuevo["limpiando"] = False  # se reactiva abajo en el finally, pero se deja explícito acá también
                ESTADO_ACTUAL.clear()
                ESTADO_ACTUAL.update(nuevo)
            # Si NO era el lote activo (otro distinto sigue corriendo),
            # no se toca el estado global -- solo se limpió el progreso
            # guardado en disco de ESTE conjunto de casos.

        _log(f"Se borró el progreso y los archivos temporales de este lote "
             f"({len(casos)} caso(s)) a pedido del usuario.", level="warn")
        return jsonify({"ok": True, "habia_progreso": existia, "estado_boton": "listo"})
    finally:
        with ESTADO_LOCK:
            ESTADO_ACTUAL["limpiando"] = False


@app.route("/api/estado")
@auth.login_requerido
def api_estado():
    """Sondeado por el navegador cada cierto tiempo mientras corre un
    lote: devuelve solo las líneas de log NUEVAS desde `offset`, más los
    contadores/estado actuales. Cuando `finished` es true, también manda
    los resultados finales.

    `estado_boton` es la ÚNICA fuente de verdad para qué debe mostrar el
    botón principal ('ejecutando' | 'pausado' | 'listo') — el frontend
    nunca lo calcula por su cuenta, así nunca queda desincronizado del
    backend (ver `_estado_boton()`)."""
    sesion_datos = auth.sesion_actual()
    estado_boton = _estado_boton(sesion_datos)
    offset = request.args.get("offset", 0, type=int)
    with ESTADO_LOCK:
        logs_nuevos = ESTADO_ACTUAL["logs"][offset:]
        payload = {
            "running": ESTADO_ACTUAL["running"],
            "finished": ESTADO_ACTUAL["finished"],
            "error": ESTADO_ACTUAL["error"],
            "stats": dict(ESTADO_ACTUAL["stats"]),
            "logs": logs_nuevos,
            "log_total": len(ESTADO_ACTUAL["logs"]),
            "lote_id": ESTADO_ACTUAL["lote_id"],
            "detenido": ESTADO_ACTUAL.get("detenido", False),
            "estado_boton": estado_boton,
            "modo": ESTADO_ACTUAL.get("modo"),
            # Listas cortas, se mandan completas en cada consulta (no
            # incrementales como el log) — el navegador simplemente
            # reemplaza el contenido de esas pestañas cada vez.
            "rutas_sugeridas": list(ESTADO_ACTUAL["rutas_sugeridas"]),
            "resumen_errores": list(ESTADO_ACTUAL["resumen_errores"]),
            "urls_consulta": list(ESTADO_ACTUAL["urls_consulta"]),
        }
        if ESTADO_ACTUAL["finished"]:
            payload["resumenes"] = ESTADO_ACTUAL["resumenes"]
            payload["pdfs"] = ESTADO_ACTUAL["pdfs"]
            payload["tiene_excel"] = ESTADO_ACTUAL["tiene_excel"]
    return jsonify(payload)


@app.route("/api/progreso-pendiente")
@auth.login_requerido
def api_progreso_pendiente():
    """Se consulta UNA VEZ al cargar la página (o recargarla): ¿hay un
    progreso sin terminar de este mismo usuario+empresa, aunque el
    navegador no tenga a mano la lista de casos ni el lote_id? Si lo
    hay, se devuelve la lista de casos original guardada en
    progreso.json (ver `crear_progreso`) para que el frontend la
    reconstruya sola y muestre 'Reanudar' de una — sin que el usuario
    tenga que volver a escribir o subir nada."""
    _sanear_estado()
    if _hilo_activo():
        # Ya hay algo corriendo (posiblemente arrancado desde otra
        # pestaña/computador) -- no tiene sentido ofrecer "reanudar" algo
        # que ya está en marcha.
        return jsonify({"hay_progreso": False})
    sesion_datos = auth.sesion_actual()
    pendiente = prog.buscar_progreso_pendiente(sesion_datos.get("usuario", "?"), sesion_datos.get("empresa", "?"))
    if not pendiente:
        return jsonify({"hay_progreso": False})
    return jsonify({
        "hay_progreso": True,
        "casos": pendiente.get("casos", []),
        "clave_lote": pendiente.get("clave_lote"),
        "resumen": prog.resumen_para_mostrar(pendiente),
    })


@app.route("/descargar/<lote_id>/pdf/<nombre_pdf>")
@auth.login_requerido
def descargar_pdf(lote_id, nombre_pdf):
    ruta = OUTPUTS_DIR / lote_id / nombre_pdf
    if not ruta.exists():
        return "Archivo no encontrado", 404
    return send_file(ruta, as_attachment=True, download_name=nombre_pdf)


@app.route("/descargar/<lote_id>/excel")
@auth.login_requerido
def descargar_excel(lote_id):
    ruta = OUTPUTS_DIR / lote_id / "Informe_resultado.xlsx"
    if not ruta.exists():
        return "Archivo no encontrado", 404
    return send_file(ruta, as_attachment=True, download_name="Informe_resultado.xlsx")


def _nombre_zip_por_empresa(lote_id):
    """Arma el nombre del .zip usando la empresa con la que se inició
    sesión: "{código} - {nombre de la IPS}" (usa la sede/nombre comercial
    si existe, si no el nombre de la empresa). Ej: "BH - INVERSIONES AZALUD".
    Si por algún motivo no hay sesión con empresa, cae al nombre genérico
    de siempre (con el id del lote)."""
    empresa_cod = session.get("empresa", "")
    empresas = session.get("empresas", [])
    nombre_ips = ""
    for e in empresas:
        if e.get("cod") == empresa_cod:
            nombre_ips = e.get("sede") or e.get("nombre") or ""
            break

    if not empresa_cod or not nombre_ips:
        return f"resultado_{lote_id}"

    base = f"{empresa_cod} - {nombre_ips}"
    # Sanitizar para nombre de archivo válido en Windows
    base = re.sub(r'[<>:"/\\|?*]', "", base).strip()
    return base or f"resultado_{lote_id}"


@app.route("/descargar/<lote_id>/zip")
@auth.login_requerido
def descargar_zip(lote_id):
    carpeta = OUTPUTS_DIR / lote_id
    if not carpeta.exists():
        return "Lote no encontrado", 404

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in carpeta.iterdir():
            zf.write(item, arcname=item.name)
    buffer.seek(0)
    return send_file(
        buffer, as_attachment=True,
        download_name=f"{_nombre_zip_por_empresa(lote_id)}.zip",
        mimetype="application/zip",
    )


# ─────────────────────────────────────────────────────────────
# Modo "Buscar por factura/caso" (API real) — plantilla, carga de
# casos y disparo del proceso. La función que de verdad trae los PDF
# desde el sistema (`_obtener_documentos_caso`) queda como un pendiente
# claramente marcado: en cuanto se comparta el endpoint real (URL +
# parámetros), solo hay que completar esa función; todo lo demás
# (checklist, cola de casos, log en vivo, progreso, Excel) ya está listo.
# ─────────────────────────────────────────────────────────────

@app.route("/api/plantilla-casos")
@auth.login_requerido
def plantilla_casos():
    if not EXCEL_OK:
        return jsonify({"error": "openpyxl no instalado"}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Casos"
    ws.append(["NoCaso", "NoFactura"])
    ws.append(["192681", "F7195994"])
    ws.append(["182225", "F7195986"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out, as_attachment=True, download_name="plantilla_casos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/cargar-casos", methods=["POST"])
@auth.login_requerido
def cargar_casos():
    """Sube un .xlsx/.csv con columnas NoCaso/NoFactura (o similares) y
    devuelve la lista de casos ya parseada, para que el frontend la
    guarde y la mande a /api/iniciar-casos.

    IMPORTANTE: la búsqueda del documento siempre se hace por NoCaso —
    NoFactura NO se usa para buscar nada, solo sirve para nombrar el PDF
    de salida (así lo pidió el cliente: "los documentos siempre se
    buscarán es por número de caso, el número de factura lo pedimos es
    porque el resultado del PDF se nombrará con el número de factura").
    """
    try:
        if "file" not in request.files:
            return jsonify({"error": "No se envió archivo"}), 400
        f = request.files["file"]
        if not f.filename:
            return jsonify({"error": "Archivo vacío"}), 400

        datos = []
        if f.filename.lower().endswith(".csv"):
            raw = f.stream.read().decode("utf-8-sig")
            lector = csv.DictReader(raw.splitlines())
            for fila in lector:
                nc = (fila.get("NoCaso") or fila.get("nocaso") or fila.get("NOCASO") or
                      fila.get("NoCuenta") or fila.get("nocuenta") or fila.get("NOCUENTA"))
                nf = (fila.get("NoFactura") or fila.get("nofactura") or
                      fila.get("NOFACTURA") or fila.get("Factura") or "")
                nc = _solo_digitos(nc)
                if nc:
                    datos.append({"NoCaso": nc, "NoFactura": str(nf).strip()})
        elif f.filename.lower().endswith(".xlsx"):
            if not EXCEL_OK:
                return jsonify({"error": "openpyxl no instalado"}), 500
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            encabezados = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

            def buscar_col(nombres):
                for nombre in nombres:
                    for i, h in enumerate(encabezados):
                        if h == nombre:
                            return i
                return None

            # Se acepta "nocuenta"/"cuenta" como alias por si alguien reusa
            # la plantilla vieja, pero el campo real ya es "NoCaso".
            col_nc = buscar_col(["nocaso", "no_caso", "no caso", "caso", "nocuenta", "no_cuenta", "no cuenta", "cuenta"])
            col_nf = buscar_col(["nofactura", "no_factura", "no factura", "factura", "nfactura"])
            if col_nc is None:
                col_nc = 0
            for fila in ws.iter_rows(min_row=2, values_only=True):
                nc = str(fila[col_nc]).strip() if col_nc is not None and col_nc < len(fila) and fila[col_nc] is not None else ""
                nf = str(fila[col_nf]).strip() if col_nf is not None and col_nf < len(fila) and fila[col_nf] is not None else ""
                nc = _solo_digitos(nc)
                if nc:
                    datos.append({"NoCaso": nc, "NoFactura": nf})
        else:
            return jsonify({"error": "Sube un archivo .xlsx o .csv"}), 400

        return jsonify({"ok": True, "count": len(datos), "casos": datos})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _decodificar_base64_a_archivo(b64_string, ruta_salida):
    """Decodifica un string base64 (agregando el padding que le falte,
    algo común cuando el dato viene de una base de datos/API) y lo guarda
    tal cual en disco — para un PDF completo, no hace falta pasar por PIL
    como sí se hacía en el proyecto de referencia con la firma del
    médico (ahí además le quitaban el fondo blanco; aquí no aplica,
    porque no es un recorte de firma, es el documento completo)."""
    faltante = len(b64_string) % 4
    if faltante:
        b64_string += "=" * (4 - faltante)
    datos = base64.b64decode(b64_string)
    ruta_salida.write_bytes(datos)
    return ruta_salida


def _procesar_respuesta_archivo_pdf(contenido, es_binario, content_type, ruta_salida):
    """Guarda en disco lo que haya devuelto la API (PDF binario directo, o
    JSON con el contenido en base64 en algún campo). Se separó de
    `_obtener_documentos_caso` para poder reusarla al probar cada ruta
    una por una."""
    if es_binario:
        ruta_salida.write_bytes(contenido)
        return ruta_salida
    if isinstance(contenido, dict):
        for nombre_campo in ("Archivo", "archivo", "Pdf", "pdf", "Base64", "base64",
                              "Contenido", "contenido", "ArchivoBase64", "Data", "data"):
            if contenido.get(nombre_campo):
                return _decodificar_base64_a_archivo(contenido[nombre_campo], ruta_salida)
        raise RuntimeError(
            f"La API respondió pero no se encontró el campo con el PDF en base64 "
            f"(claves recibidas: {list(contenido.keys())})."
        )
    raise RuntimeError(
        f"Respuesta inesperada (content-type: {content_type}, tipo: {type(contenido).__name__})."
    )


def _obtener_documentos_caso(caso, sesion_datos, tipos_documento, rutas, clave_lote=None):
    """Trae el archivo real de un caso usando la API dedicada
    `obtener-archivo-pdf`. Los parámetros de conexión (ip, puerto,
    usuario, clave, empresa) son los mismos de la sesión activa; lo único
    que cambia por búsqueda son `NoCaso` y las rutas marcadas.

    MUY IMPORTANTE (aclarado por el cliente): la búsqueda del documento
    SIEMPRE se hace por `caso["NoCaso"]` — nunca por `caso["NoFactura"]`.

    Se prueba RUTA POR RUTA (no todas juntas en una sola llamada) — así,
    apenas una funciona, se sabe exactamente CUÁL fue (se registra en el
    log con su nombre), en vez de solo saber que "alguna de las que se
    marcaron" tenía el archivo. Esto es lo que permite, con el tiempo,
    saber cuál ruta dejar marcada por defecto para cada tipo de caso.
    """
    no_caso = str(caso.get("NoCaso", "")).strip()
    if not rutas:
        raise RuntimeError(
            "No se marcó ninguna ruta/centro de digitalización en el checklist — "
            "sin al menos una ruta no hay dónde buscar los documentos del caso."
        )

    carpeta_temp = UPLOADS_DIR / "_casos_temp"
    carpeta_temp.mkdir(parents=True, exist_ok=True)

    def _probar_una_ruta(item):
        idx, ruta_info = item
        ruta_red = ruta_info.get("ruta", "")
        nombre_ruta = ruta_info.get("nombre", "(sin nombre)")
        if not ruta_red:
            return None
        _log(f"Caso {no_caso}: probando en la ruta \"{nombre_ruta}\"...")
        # Se registra la URL EXACTA antes de llamar, así queda disponible
        # en la pestaña "URLs de búsqueda" incluso si la llamada falla o
        # tarda -- es lo que permite copiarla y probarla a mano en el
        # navegador cuando da 404.
        url_debug = auth.construir_url_archivo_pdf(
            sesion_datos["ip"], sesion_datos["puerto"], sesion_datos["usuario"],
            sesion_datos["password"], sesion_datos["empresa"], no_caso, [ruta_red],
        )
        _registrar_url_consulta(no_caso, nombre_ruta, url_debug)
        try:
            contenido, es_binario, content_type = auth.get_archivo_pdf_api(
                sesion_datos["ip"], sesion_datos["puerto"], sesion_datos["usuario"],
                sesion_datos["password"], sesion_datos["empresa"], no_caso, [ruta_red],
            )
            ruta_salida = carpeta_temp / f"{no_caso}_{idx}_{datetime.now().strftime('%H%M%S%f')}.pdf"
            resultado = _procesar_respuesta_archivo_pdf(contenido, es_binario, content_type, ruta_salida)
            _log(f"Caso {no_caso}: ENCONTRADO en la ruta \"{nombre_ruta}\"", level="ok")
            _registrar_ruta_sugerida(no_caso, nombre_ruta)
            if clave_lote:
                # Se marca de una que este caso YA trajo al menos un
                # archivo real de la API -- si el proceso se corta justo
                # después (antes de terminar de clasificar/combinar),
                # una reanudación queda con registro explícito de que la
                # descarga sí se alcanzó a hacer.
                prog.registrar_archivo_descargado(clave_lote, no_caso)
            return ("ok", resultado, nombre_ruta)
        except Exception as e:
            _log(f"Caso {no_caso}: no está en \"{nombre_ruta}\" ({e})", level="warn")
            return ("error", f"{nombre_ruta}: {e}", None)

    # Las rutas de un mismo caso se prueban EN PARALELO (son llamadas de
    # red, no trabajo de CPU — no compiten por el procesador entre sí) en
    # vez de una por una como antes. Esto no cambia el resultado (mismos
    # archivos, misma combinación), solo el tiempo: si un caso tiene 3-4
    # rutas marcadas, antes eran 3-4 esperas seguidas, ahora es una sola
    # espera (la más lenta de todas). Se limita a MAX_RUTAS_PARALELAS a
    # la vez para no mandarle de golpe muchas conexiones simultáneas al
    # servidor de archivos del cliente.
    MAX_RUTAS_PARALELAS = 4
    rutas_validas = [(i, r) for i, r in enumerate(rutas) if r.get("ruta")]
    encontrados = []  # (ruta_salida, nombre_ruta) de cada ruta que SÍ tuvo algo
    errores = []
    if rutas_validas:
        with ThreadPoolExecutor(max_workers=min(len(rutas_validas), MAX_RUTAS_PARALELAS)) as ex:
            for resultado in ex.map(_probar_una_ruta, rutas_validas):
                if resultado is None:
                    continue
                tipo, valor, nombre_ruta = resultado
                if tipo == "ok":
                    encontrados.append((valor, nombre_ruta))
                else:
                    errores.append(valor)

    if not encontrados:
        raise RuntimeError(
            f"No se encontró el caso {no_caso} en ninguna de las {len(rutas)} ruta(s) probadas. "
            f"Detalle por ruta: " + " | ".join(errores)
        )

    if len(encontrados) == 1:
        # Caso simple: solo una ruta tenía algo, se usa tal cual.
        return encontrados[0][0]

    # Varias rutas tenían archivo para este caso (ej. la cédula está en
    # "Admisiones" y el SIRAS en "SIRAS Campbell") — se combinan todas en
    # un solo PDF antes de clasificar, porque clasificador.py ya sabe
    # encontrar SIRAS/Cédula/Tarjeta sin importar en qué página estén.
    nombres_combinados = ", ".join(nombre for _, nombre in encontrados)
    _log(f"Caso {no_caso}: se encontró en {len(encontrados)} rutas distintas "
         f"({nombres_combinados}) — se combinan en un solo PDF antes de clasificar.", level="info")
    ruta_combinada = carpeta_temp / f"{no_caso}_combinado_{datetime.now().strftime('%H%M%S%f')}.pdf"
    doc_combinado = fitz.open()
    for ruta_pdf, _ in encontrados:
        with fitz.open(str(ruta_pdf)) as doc_parcial:
            doc_combinado.insert_pdf(doc_parcial)
    doc_combinado.save(str(ruta_combinada))
    doc_combinado.close()
    return ruta_combinada


def _procesar_un_caso(item, sesion_datos, tipos_documento, rutas, carpeta_lote, clave_lote=None):
    caso = item
    no_caso = str(caso.get("NoCaso", "")).strip()
    no_factura = str(caso.get("NoFactura", "")).strip()
    # Solo para mostrar en el log/tabla mientras no hay PDF real todavía;
    # el nombre final del PDF de salida lo pone clasificador.py usando el
    # número de factura que trae el propio documento, no este valor.
    nombre_referencia = no_factura or no_caso
    rutas_nombres = ", ".join(r.get("nombre", "?") for r in rutas) if rutas else "(ninguna)"
    _log(f"Buscando documentos del caso {no_caso}... (usuario: {sesion_datos.get('usuario','?')}, "
         f"empresa: {sesion_datos.get('empresa','?')}, rutas marcadas: {rutas_nombres})")
    if clave_lote:
        prog.marcar_caso_actual(clave_lote, no_caso)
    try:
        ruta_pdf = _obtener_documentos_caso(caso, sesion_datos, tipos_documento, rutas, clave_lote=clave_lote)
        # El checklist de "Documentos del paciente"/"Documentos del caso"
        # decide QUÉ traer del archivo encontrado — si el usuario solo
        # marcó Cédula y Tarjeta (sin SIRAS), el PDF final no debe incluir
        # SIRAS aunque el archivo lo traiga. Se traduce cada nombre
        # marcado a la categoría real que el bot identifica (ver
        # checklist_config.categoria_para_nombre); los tipos marcados que
        # todavía son "(P)" (no implementados) no producen ninguna
        # categoría y simplemente no filtran nada extra.
        categorias_permitidas = {
            cat for doc in tipos_documento
            if (cat := chk.categoria_para_nombre(doc.get("nombre", "")))
        } or None  # None = no reciben ningún tipo reconocido, no se filtra nada (se incluye todo lo detectado)
        resumen, generado = _procesar_un_archivo(
            (f"{nombre_referencia}.pdf", ruta_pdf), categorias_permitidas=categorias_permitidas)
        if clave_lote:
            # Se persiste INMEDIATAMENTE, no se espera a que termine el
            # lote entero -- si el proceso se corta un segundo después,
            # este registro ya queda a salvo y una reanudación no lo
            # vuelve a tocar.
            prog.marcar_registro_completado(clave_lote, no_caso, resumen, generado)
        return resumen, generado
    except (NotImplementedError, RuntimeError) as e:
        _log(f"Caso {no_caso}: {e}", level="error")
        _sumar_stat("err")
        _registrar_error_resumen(no_caso, e)
        resumen = {
            "Archivo original": nombre_referencia,
            "No. Factura": no_factura,
            "SIRAS": "?", "Pág. SIRAS": "",
            "Cédula/Migrante": "?", "Pág. Cédula": "",
            "Tarjeta Propiedad": "?", "Pág. Tarjeta": "",
            "Estado": "ERROR",
            "Observaciones": str(e),
            "Archivo generado": "",
        }
        if clave_lote:
            prog.marcar_registro_error(clave_lote, no_caso, str(e), resumen)
        return resumen, None
    except Exception as e:
        traceback.print_exc()
        _log(f"Caso {no_caso}: error inesperado — {e}", level="error")
        _sumar_stat("err")
        _registrar_error_resumen(no_caso, f"Error inesperado: {e}")
        resumen = {
            "Archivo original": nombre_referencia,
            "No. Factura": no_factura,
            "SIRAS": "?", "Pág. SIRAS": "",
            "Cédula/Migrante": "?", "Pág. Cédula": "",
            "Tarjeta Propiedad": "?", "Pág. Tarjeta": "",
            "Estado": "ERROR",
            "Observaciones": f"Error inesperado: {e}",
            "Archivo generado": "",
        }
        if clave_lote:
            prog.marcar_registro_error(clave_lote, no_caso, f"Error inesperado: {e}", resumen)
        return resumen, None


def _ejecutar_casos_en_hilo(clave_lote, casos, sesion_datos, tipos_documento, rutas):
    """`clave_lote` es la clave ESTABLE del lote (ver progreso.py) — se
    usa como nombre de carpeta tanto para uploads temporales como para
    outputs, así una reanudación reutiliza la misma carpeta de salida
    en vez de generar una nueva y perder los PDF ya generados antes."""
    lote_id = clave_lote  # mismo valor: rutas de /descargar/<lote_id>/... siguen funcionando igual
    carpeta_lote = UPLOADS_DIR / lote_id
    carpeta_lote.mkdir(parents=True, exist_ok=True)
    carpeta_salida = OUTPUTS_DIR / lote_id
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    pendientes, progreso_previo = prog.registros_pendientes(clave_lote, casos)
    ya_completados = len(casos) - len(pendientes)
    if progreso_previo and ya_completados:
        _log(f"Reanudando lote — {ya_completados} de {len(casos)} caso(s) ya estaban "
             f"resueltos en un intento anterior, no se vuelven a procesar.", level="info")
    _log(f"Iniciando búsqueda de {len(pendientes)} caso(s) pendiente(s) "
         f"(de {len(casos)} en total)...")

    error_general = None
    detenido = False

    try:
        ex = ThreadPoolExecutor(max_workers=HILOS_ARCHIVOS)
        futuros = {
            ex.submit(_procesar_un_caso, caso, sesion_datos, tipos_documento, rutas,
                      carpeta_lote, clave_lote): idx
            for idx, caso in enumerate(pendientes)
        }
        try:
            for futuro in as_completed(futuros):
                _, generado = futuro.result()
                if generado:
                    # Se escribe a disco DE UNA, no se espera al final del
                    # lote -- si el proceso se corta un segundo después,
                    # este PDF ya quedó guardado y no se pierde.
                    nombre_pdf, contenido = generado
                    with open(carpeta_salida / nombre_pdf, "wb") as fh:
                        fh.write(contenido)
                if EVENTO_DETENER.is_set():
                    detenido = True
                    _log("Detención solicitada — se completan los casos que ya estaban en "
                         "proceso y se arma el resultado con lo que haya listo hasta ahora.", level="warn")
                    break
        finally:
            ex.shutdown(wait=True, cancel_futures=True)
    except Exception as e:
        traceback.print_exc()
        error_general = str(e)

    # El resultado final se arma SIEMPRE a partir de progreso.json (no de
    # variables en memoria de esta corrida) -- así, tanto si el lote es
    # nuevo como si es una reanudación, el Excel/ZIP quedan con TODO lo
    # resuelto hasta ahora (lo de antes + lo de esta corrida), en el
    # mismo orden en que se enviaron los casos originalmente.
    resumenes_por_caso = dict(prog.resumenes_ya_completados(clave_lote))
    resumenes = [
        resumenes_por_caso[str(c.get("NoCaso", "")).strip()]
        for c in casos
        if str(c.get("NoCaso", "")).strip() in resumenes_por_caso
    ]

    excel_bytes = None
    if EXCEL_OK and resumenes:
        try:
            excel_bytes = generar_excel(resumenes)
            with open(carpeta_salida / "Informe_resultado.xlsx", "wb") as fh:
                fh.write(excel_bytes)
        except Exception as e:
            _log(f"No se pudo generar el Excel: {e}", level="warn")

    # La carpeta de uploads TEMPORAL sí se limpia siempre (descargas de
    # trabajo, ya no hacen falta) -- la de outputs (PDFs + Excel) se deja,
    # es la que ve/descarga el usuario y la que permite reanudar sin
    # perder lo ya generado.
    shutil.rmtree(carpeta_lote, ignore_errors=True)

    if detenido:
        prog.marcar_estado(clave_lote, "detenido")
    elif error_general:
        prog.marcar_estado(clave_lote, "error")
    else:
        prog.marcar_estado(clave_lote, "completado")

    if detenido:
        _log(f"Búsqueda detenida por el usuario — se procesaron {len(resumenes)} de {len(casos)} "
             f"caso(s) antes de parar. El ZIP/Excel ya están listos con eso.", level="warn")
    else:
        _log("Búsqueda terminada." if not error_general else "Búsqueda terminada con errores.",
             level="ok" if not error_general else "error")

    progreso_final = prog.cargar_progreso(clave_lote)
    pdfs_finales = (progreso_final or {}).get("documentos_generados", [])

    with ESTADO_LOCK:
        ESTADO_ACTUAL["running"] = False
        ESTADO_ACTUAL["finished"] = True
        ESTADO_ACTUAL["error"] = error_general
        ESTADO_ACTUAL["detenido"] = detenido
        ESTADO_ACTUAL["resumenes"] = resumenes
        ESTADO_ACTUAL["pdfs"] = pdfs_finales
        ESTADO_ACTUAL["tiene_excel"] = bool(excel_bytes)


@app.route("/api/iniciar-casos", methods=["POST"])
@auth.login_requerido
def iniciar_casos():
    _sanear_estado()
    with ESTADO_LOCK:
        if _hilo_vivo_sin_lock():
            return jsonify({"ok": False, "error": "Ya hay un lote en proceso, espera a que termine."}), 409
        ESTADO_ACTUAL["running"] = True  # reservar el turno de una (ver nota en /procesar)

    EVENTO_DETENER.clear()
    data = request.get_json() or {}
    casos_crudos = data.get("casos", [])
    tipos_documento = data.get("tipos_documento", [])
    rutas = data.get("rutas", [])
    # 'decision' la manda el navegador SOLO en el segundo intento, después
    # de que el usuario elige qué hacer con un progreso previo encontrado:
    # "reanudar" (seguir donde quedó) o "limpiar" (empezar de cero).
    decision = data.get("decision")

    # Normalización defensiva de NoCaso (por si el navegador mandó algo
    # sin limpiar): se quita todo lo que no sea dígito -- "641-757",
    # "641 757", "N° 641757", etc. todos terminan siendo "641757". Esto
    # es la MISMA normalización que ya se aplica al leer el Excel/CSV
    # (ver `_solo_digitos` y `cargar_casos`); se repite acá para cubrir
    # también la carga manual desde el navegador.
    casos = []
    descartados = 0
    for c in casos_crudos:
        nc = _solo_digitos(c.get("NoCaso"))
        if not nc:
            descartados += 1
            continue
        casos.append({"NoCaso": nc, "NoFactura": str(c.get("NoFactura", "")).strip()})
    if descartados:
        _log(f"Se descartaron {descartados} caso(s) sin ningún dígito en el número de caso "
             f"(no hay forma de buscarlos).", level="warn")

    if not casos:
        with ESTADO_LOCK:
            ESTADO_ACTUAL["running"] = False
        return jsonify({"ok": False, "error": "No hay casos válidos para procesar."}), 400

    sesion_datos = auth.sesion_actual()
    clave_lote = prog.calcular_clave_lote(casos, sesion_datos.get("usuario", "?"), sesion_datos.get("empresa", "?"))

    # ¿Ya existe progreso sin terminar para este MISMO conjunto de casos
    # (mismo usuario/empresa)? Si el navegador todavía no mandó una
    # decisión, se le devuelve el resumen para que el usuario elija —
    # nunca se reanuda ni se limpia nada en silencio.
    if decision is None and prog.existe_progreso_activo(clave_lote):
        with ESTADO_LOCK:
            ESTADO_ACTUAL["running"] = False
        progreso_previo = prog.cargar_progreso(clave_lote)
        return jsonify({
            "ok": False,
            "necesita_decision": True,
            "progreso": prog.resumen_para_mostrar(progreso_previo),
        })

    if decision == "limpiar":
        _log("Limpiando progreso anterior de este mismo lote a pedido del usuario...", level="warn")
        prog.limpiar_todo(clave_lote, carpetas_extra=[UPLOADS_DIR / clave_lote, OUTPUTS_DIR / clave_lote])

    if not prog.cargar_progreso(clave_lote):
        prog.crear_progreso(clave_lote, casos, sesion_datos)
    elif decision == "reanudar":
        _log("Reanudando lote donde había quedado...", level="info")

    with ESTADO_LOCK:
        ESTADO_ACTUAL.update({
            "running": True,
            "finished": False,
            "error": None,
            "logs": [],
            "stats": {"total": len(casos), "ok": 0, "err": 0},
            "lote_id": clave_lote,
            "resumenes": [],
            "pdfs": [],
            "tiene_excel": False,
            "rutas_sugeridas": [],
            "resumen_errores": [],
            "urls_consulta": [],
            "modo": "casos",
        })

    hilo = threading.Thread(
        target=_ejecutar_casos_en_hilo,
        args=(clave_lote, casos, sesion_datos, tipos_documento, rutas),
        daemon=True,
    )
    with ESTADO_LOCK:
        ESTADO_ACTUAL["hilo"] = hilo
    hilo.start()

    return jsonify({"ok": True, "lote_id": clave_lote, "total": len(casos), "estado_boton": "ejecutando"})


def _abrir_navegador_diferido(url, espera=1.5):
    """Espera un momento (a que el servidor ya esté escuchando) y abre el
    navegador solo, para que el usuario no tenga que copiar la URL a mano
    ni tocar la ventana de la consola para nada."""
    import webbrowser
    time.sleep(espera)
    try:
        webbrowser.open(url)
    except Exception:
        pass  # si falla, el usuario igual puede abrir la URL a mano


if __name__ == "__main__":
    # PORT lo pone Railway (y casi cualquier plataforma en la nube)
    # automáticamente -- en la máquina local, si no existe, se usa el
    # puerto de siempre (5057) para no romper iniciar.bat.
    puerto = int(os.environ.get("PORT", 5057))
    # RAILWAY_ENVIRONMENT (o cualquier variable que Railway defina) es la
    # señal de que esto NO es una laptop con pantalla -- ahí no tiene
    # sentido (ni funcionaría) intentar abrir un navegador. En producción
    # real esto ni se ejecuta: Railway arranca con gunicorn (ver
    # railway.json), que IMPORTA este archivo como módulo en vez de
    # correrlo como script, así que este bloque completo queda sin usar.
    # Este chequeo es solo un respaldo por si alguien corre
    # `python app.py` directo dentro del contenedor de la nube.
    es_nube = bool(os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("PORT"))
    if es_nube:
        print("=" * 60)
        print(" Extractor de Documentos Digitales — arrancando en modo servidor")
        print(f" Escuchando en el puerto {puerto}")
        print("=" * 60)
        app.run(host="0.0.0.0", port=puerto, debug=False)
    else:
        URL_LOCAL = f"http://127.0.0.1:{puerto}"
        print("=" * 60)
        print(" Extractor de Documentos Digitales — v1 (pruebas locales)")
        print(f" Abriendo el navegador en: {URL_LOCAL}")
        print(" (si no se abre solo, entra a esa dirección a mano)")
        print("=" * 60)
        threading.Thread(target=_abrir_navegador_diferido, args=(URL_LOCAL,), daemon=True).start()
        app.run(host="0.0.0.0", port=puerto, debug=False)
